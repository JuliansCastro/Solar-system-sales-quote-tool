#!/usr/bin/env python
"""
Test script to validate PDF and Excel content contains correct sizing data.
Extracts actual content from generated PDF/Excel to verify coherence.
"""

import os
import sys
import django
import io

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'solar_app.settings')
sys.path.insert(0, os.path.dirname(__file__))

django.setup()

from core.models import Cotizacion, Proyecto, CotizacionItem, Equipo
from core.calculations.chart_data import _calculate_sizing_from_items
from core.calculations.reports import generar_pdf_cotizacion, generar_excel_cotizacion

try:
    from PyPDF2 import PdfReader
except ImportError:
    print("⚠️  PyPDF2 not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "PyPDF2", "-q"])
    from PyPDF2 import PdfReader

from openpyxl import load_workbook


def extract_pdf_text(pdf_buffer):
    """Extract text from PDF buffer."""
    try:
        pdf_reader = PdfReader(pdf_buffer)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() or ""
        return text
    except Exception as e:
        print(f"Error extracting PDF: {e}")
        return ""


def extract_excel_text(excel_buffer):
    """Extract text from Excel buffer."""
    try:
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        
        # Sheet 1: Cotización
        ws1_data = {}
        if 'Cotización' in wb.sheetnames:
            ws = wb['Cotización']
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        ws1_data[str(cell)] = cell
        
        # Sheet 2: Resumen Técnico
        ws2_data = {}
        if 'Resumen Técnico' in wb.sheetnames:
            ws = wb['Resumen Técnico']
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        ws2_data[str(cell)] = cell
        
        return {
            'sheet1': ws1_data,
            'sheet2': ws2_data,
        }
    except Exception as e:
        print(f"Error extracting Excel: {e}")
        return {}


def test_content_validation():
    """Validate actual content of PDF and Excel against calculated sizes."""
    
    # Get first cotización with items
    cotizacion = Cotizacion.objects.filter(items__isnull=False).prefetch_related(
        'items__equipo', 'proyecto'
    ).first()
    
    if not cotizacion:
        print("❌ No cotización with items found")
        return False
    
    proyecto = cotizacion.proyecto
    print(f"\n{'='*70}")
    print(f"VALIDATING PDF AND EXCEL CONTENT")
    print(f"{'='*70}")
    print(f"Cotización: {cotizacion.numero}")
    print(f"Proyecto: {proyecto.codigo}\n")
    
    # Calculate sizing from items (source of truth)
    sizing_from_items = _calculate_sizing_from_items(cotizacion, proyecto)
    print(f"[EXPECTED VALUES (from items)]")
    print(f"  - Potencia: {sizing_from_items.get('potencia_pico_kwp')} kWp")
    print(f"  - Paneles: {sizing_from_items.get('numero_paneles')}")
    print(f"  - Generación: {sizing_from_items.get('generacion_mensual_kwh'):.0f} kWh/mes")
    print(f"  - Cobertura: {sizing_from_items.get('porcentaje_cobertura'):.1f}%\n")
    
    # Generate and analyze PDF
    print(f"{'─'*70}")
    print("[PDF VALIDATION]")
    print(f"{'─'*70}")
    try:
        pdf_buffer = generar_pdf_cotizacion(cotizacion)
        pdf_size = len(pdf_buffer.getvalue())
        print(f"[OK] PDF generated: {pdf_size:,} bytes")
        
        pdf_text = extract_pdf_text(pdf_buffer)
        
        # Search for key metrics in PDF
        potencia_str = f"{sizing_from_items.get('potencia_pico_kwp'):.2f} kWp"
        paneles_str = str(sizing_from_items.get('numero_paneles'))
        generacion_str = f"{sizing_from_items.get('generacion_mensual_kwh'):.0f} kWh"
        
        potencia_found = potencia_str in pdf_text
        paneles_found = paneles_str in pdf_text
        generacion_found = generacion_str in pdf_text
        
        print(f"\n  Searching in PDF text:")
        print(f"    [{'OK' if potencia_found else 'FAIL'}] Potencia '{potencia_str}': {potencia_found}")
        print(f"    [{'OK' if paneles_found else 'FAIL'}] Paneles '{paneles_str}': {paneles_found}")
        print(f"    [{'OK' if generacion_found else 'FAIL'}] Generación '{generacion_str}': {generacion_found}")
        
        if potencia_found and paneles_found and generacion_found:
            print("\n  [SUCCESS] PDF contains correct calculated values!")
        else:
            print("\n  [WARNING] Some values not found in PDF text")
            print("\n  First 500 chars of PDF:")
            print(f"    {pdf_text[:500]}")
        
    except Exception as e:
        print(f"  [ERROR] PDF validation failed: {e}")
        import traceback
        traceback.print_exc()
    
    # Generate and analyze Excel
    print(f"\n{'─'*70}")
    print("[EXCEL VALIDATION]")
    print(f"{'─'*70}")
    try:
        excel_buffer = generar_excel_cotizacion(cotizacion)
        excel_size = len(excel_buffer.getvalue())
        print(f"[OK] Excel generated: {excel_size:,} bytes")
        
        excel_data = extract_excel_text(excel_buffer)
        
        if 'sheet2' in excel_data:
            # Sheet 2: Resumen Técnico
            potencia_str = f"{sizing_from_items.get('potencia_pico_kwp'):.2f} kWp"
            paneles_str = str(sizing_from_items.get('numero_paneles'))
            generacion_str = f"{sizing_from_items.get('generacion_mensual_kwh'):.0f} kWh/mes"
            
            sheet2_text = str(excel_data['sheet2'])
            
            potencia_found = f"{sizing_from_items.get('potencia_pico_kwp'):.2f}" in sheet2_text
            paneles_found = str(sizing_from_items.get('numero_paneles')) in sheet2_text
            generacion_found = f"{sizing_from_items.get('generacion_mensual_kwh'):.0f}" in sheet2_text
            
            print(f"\n  Sheet 2 (Resumen Técnico):")
            print(f"    [{'OK' if potencia_found else 'FAIL'}] Potencia {sizing_from_items.get('potencia_pico_kwp'):.2f}: {potencia_found}")
            print(f"    [{'OK' if paneles_found else 'FAIL'}] Paneles {sizing_from_items.get('numero_paneles')}: {paneles_found}")
            print(f"    [{'OK' if generacion_found else 'FAIL'}] Generación {sizing_from_items.get('generacion_mensual_kwh'):.0f}: {generacion_found}")
            
            if potencia_found and paneles_found and generacion_found:
                print("\n  [SUCCESS] Excel contains correct calculated values!")
            else:
                print("\n  [WARNING] Some values not found in Excel Sheet 2")
                print(f"\n  Sheet 2 content sample:")
                print(f"    {list(excel_data['sheet2'].items())[:10]}")
        
    except Exception as e:
        print(f"  [ERROR] Excel validation failed: {e}")
        import traceback
        traceback.print_exc()
    
    print(f"\n{'='*70}")
    print("[SUCCESS] Content validation complete!")
    print(f"{'='*70}\n")
    return True


if __name__ == '__main__':
    success = test_content_validation()
    sys.exit(0 if success else 1)

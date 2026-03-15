"""
Reports module – PDF & Excel generation for quotes.
Uses ReportLab for PDF generation and openpyxl for Excel.
"""

import io
import locale
import logging
from datetime import datetime
from decimal import Decimal

from django.conf import settings

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image, HRFlowable, PageBreak,
)
from reportlab.graphics.shapes import Drawing, Rect, String, Line
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.textlabels import Label
from reportlab.graphics.widgets.markers import makeMarker
from plotly import graph_objects as go
from plotly.io import to_image

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers,
)
from openpyxl.utils import get_column_letter

from .chart_data import build_cotizacion_charts_payload


logger = logging.getLogger(__name__)


# ═════════════════════════════════════════════
# HELPERS
# ═════════════════════════════════════════════

def _fmt_cop(value):
    """Format a number as Colombian Pesos."""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return "$0"
    if v >= 0:
        return f"${v:,.0f}".replace(",", ".")
    return f"-${abs(v):,.0f}".replace(",", ".")


def _company_info():
    """Retrieve company info from DB (CompanySettings) with fallback to settings."""
    from .models import CompanySettings
    try:
        obj = CompanySettings.load()
        return {
            'name': obj.name or getattr(settings, 'COMPANY_NAME', 'Solar Energy Solutions'),
            'phone': obj.phone or getattr(settings, 'COMPANY_PHONE', ''),
            'email': obj.email or getattr(settings, 'COMPANY_EMAIL', ''),
            'address': obj.address or getattr(settings, 'COMPANY_ADDRESS', ''),
            'nit': obj.nit or getattr(settings, 'COMPANY_NIT', ''),
            'logo': obj.logo.path if obj.logo else getattr(settings, 'COMPANY_LOGO', ''),
        }
    except Exception:
        return {
            'name': getattr(settings, 'COMPANY_NAME', 'Solar Energy Solutions'),
            'phone': getattr(settings, 'COMPANY_PHONE', ''),
            'email': getattr(settings, 'COMPANY_EMAIL', ''),
            'address': getattr(settings, 'COMPANY_ADDRESS', ''),
            'nit': getattr(settings, 'COMPANY_NIT', ''),
            'logo': getattr(settings, 'COMPANY_LOGO', ''),
        }


# ═══════════════════════════════════════════════════════
#  CUSTOM STYLES
# ═══════════════════════════════════════════════════════

SOLAR_AMBER = colors.HexColor('#f59e0b')
SOLAR_DARK = colors.HexColor('#92400e')
SOLAR_GREEN = colors.HexColor('#166534')
CORP_DARK = colors.HexColor('#1f2937')
CORP_MUTED = colors.HexColor('#6b7280')
HEADER_BG = colors.HexColor('#f59e0b')
HEADER_TEXT = colors.white
ROW_ALT = colors.HexColor('#fffbeb')
BORDER_COLOR = colors.HexColor('#e5e7eb')


def _get_styles():
    """Build custom paragraph styles for the PDF."""
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        'CompanyName',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=SOLAR_DARK,
        spaceAfter=2,
    ))
    styles.add(ParagraphStyle(
        'DocTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=SOLAR_DARK,
        alignment=TA_RIGHT,
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading3'],
        fontSize=11,
        textColor=SOLAR_DARK,
        spaceBefore=14,
        spaceAfter=6,
        borderPadding=(0, 0, 4, 0),
    ))
    styles.add(ParagraphStyle(
        'SmallText',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#6b7280'),
    ))
    styles.add(ParagraphStyle(
        'ConditionsText',
        parent=styles['Normal'],
        fontSize=8,
        leading=11,
        textColor=colors.HexColor('#4b5563'),
    ))
    styles.add(ParagraphStyle(
        'ExplainText',
        parent=styles['Normal'],
        fontSize=8,
        leading=12,
        textColor=colors.HexColor('#374151'),
        spaceAfter=4,
        alignment=TA_JUSTIFY,
    ))
    styles.add(ParagraphStyle(
        'RightAligned',
        parent=styles['Normal'],
        alignment=TA_RIGHT,
        fontSize=9,
    ))
    styles.add(ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
    ))
    styles.add(ParagraphStyle(
        'ExecutiveTitle',
        parent=styles['Heading3'],
        fontSize=10,
        textColor=CORP_DARK,
        alignment=TA_LEFT,
        spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        'ExecutiveMetric',
        parent=styles['Normal'],
        fontSize=12,
        leading=14,
        textColor=CORP_DARK,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        'ExecutiveLabel',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
        textColor=CORP_MUTED,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        'InventoryNote',
        parent=styles['Normal'],
        fontSize=8,
        leading=11,
        textColor=CORP_DARK,
        alignment=TA_JUSTIFY,
    ))
    return styles


# ═══════════════════════════════════════════════════════
#  PDF GENERATION
# ═══════════════════════════════════════════════════════

def generar_pdf_cotizacion(cotizacion):
    """
    Generate a professional PDF for a cotización.
    Returns an io.BytesIO buffer.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        topMargin=0.6 * inch,
        bottomMargin=0.6 * inch,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        title=f"Cotización {cotizacion.numero}",
    )
    styles = _get_styles()
    story = []
    company = _company_info()
    proyecto = cotizacion.proyecto
    cliente = proyecto.cliente
    chart_payload = None
    requires_charts = (not cotizacion.usar_total_manual) or bool(proyecto.potencia_pico_kwp)
    if requires_charts:
        chart_payload = build_cotizacion_charts_payload(cotizacion)

    # ─── Header ───
    story.append(_build_header(cotizacion, company, styles))
    story.append(Spacer(1, 12))
    story.append(HRFlowable(
        width="100%", thickness=2, color=SOLAR_AMBER,
        spaceAfter=12, spaceBefore=0,
    ))

    # ─── Client & Project Info ───
    story.append(_build_info_section(cotizacion, cliente, proyecto, styles))
    story.append(Spacer(1, 10))

    # ─── Items Table ───
    if not cotizacion.usar_total_manual:
        story.append(Paragraph("Detalle de Equipos y Servicios", styles['SectionTitle']))
        story.append(_build_items_table(cotizacion, styles))
        story.append(Spacer(1, 8))

    # ─── Totals ───
    story.append(_build_totals_table(cotizacion, styles))
    story.append(Spacer(1, 14))

    # ─── Executive Summary (persuasive, commercial) ───
    story.append(_build_executive_summary(cotizacion, proyecto, styles))
    story.append(Spacer(1, 10))

    # ─── Inventory Selection Rationale ───
    if not cotizacion.usar_total_manual:
        story.append(_build_equipment_selection_note(cotizacion, proyecto, styles))
        story.append(Spacer(1, 10))

    # ─── Project Sizing Summary ───
    if proyecto.potencia_pico_kwp:
        story.append(PageBreak())
        story.append(Paragraph("Resumen Técnico del Sistema", styles['SectionTitle']))
        story.append(_build_sizing_summary(proyecto, styles))
        story.append(Spacer(1, 10))

    # ─── Cost Distribution Chart ───
    if not cotizacion.usar_total_manual and chart_payload:
        dist = chart_payload.get('costo_por_componente', {})
        if dist.get('labels') and dist.get('values'):
            story.append(Paragraph("Distribución de Costos", styles['SectionTitle']))
            story.append(_build_cost_pie_chart(chart_payload, styles))
            story.append(Spacer(1, 10))

    # ─── Dimensionamiento Charts (page break for better layout) ───
    if proyecto.potencia_pico_kwp:
        story.append(PageBreak())
        story.append(Paragraph("Análisis de Dimensionamiento", styles['SectionTitle']))
        story.append(HRFlowable(
            width="100%", thickness=1, color=SOLAR_AMBER,
            spaceAfter=8, spaceBefore=0,
        ))

        # Consumption vs Generation
        consumo_data = (chart_payload or {}).get('consumo_comparacion', {})
        consumo = float(consumo_data.get('actual', 0))
        generacion = float(consumo_data.get('generacion', 0))
        if consumo > 0:
            story.append(_build_consumption_bar_chart(chart_payload, styles))
            story.append(Spacer(1, 8))
            cobertura = (generacion / consumo * 100) if consumo > 0 else 0
            ahorro_kwh = generacion if generacion <= consumo else consumo
            texto_consumo = (
                f"<b>¿Cuál es su consumo Vs el que vá generar?</b> Su consumo mensual es de <b>{consumo:.0f} kWh</b>, y nuestro sistema "
                f"solar generará <b>{generacion:.0f} kWh</b> al mes, cubriendo el <b>{cobertura:.1f}%</b> de su demanda energética. "
                f"Esto significa <b>{ahorro_kwh:.0f} kWh de energía limpia</b> producidos directamente en su propiedad, otorgándole "
                f"<b>independencia energética inmediata</b> y una <b>reducción dramática en su factura de electricidad</b> desde el primer mes. "
                f"Produce su propia energía limpia, segura y renovable, liberándose de la volatilidad de los precios."
            )
            story.append(Paragraph(texto_consumo, styles['ExplainText']))
            story.append(Spacer(1, 10))

        # Financial Projection (ROI)
        ahorro_anual = float(proyecto.ahorro_mensual or 0) * 12
        costo_total = float(proyecto.costo_total or 0)
        if ahorro_anual > 0 and costo_total > 0:
            story.append(_build_roi_chart(chart_payload, styles))
            story.append(Spacer(1, 8))
            roi_anos = costo_total / ahorro_anual if ahorro_anual > 0 else 0
            ahorro_25_anos = ahorro_anual * 25
            texto_roi = (
                f"<b>¿Por qué es importante esta proyección?</b> Su inversión de <b>${costo_total:,.0f}</b> se recuperará en tan solo "
                f"<b>{roi_anos:.1f} años</b> con un ahorro anual de <b>${ahorro_anual:,.0f}</b>. A los 25 años de operación, habrá ahorrado "
                f"<b>${ahorro_25_anos:,.0f}</b>, convirtiendo esto en una <b>inversión que genera beneficios crecientes</b> durante más de 25 años. "
                f"Año tras año, mientras otros pagan facturas cada vez más altas, usted ahorra dinero e incrementa su patrimonio. "
                f"Representa <b>seguridad financiera</b> a largo plazo, blindándose contra futuros aumentos en tarifas."
            )
            story.append(Paragraph(texto_roi, styles['ExplainText']))
            story.append(Spacer(1, 10))

        # PVGIS Radiation & HSP Charts
        radiacion_mensual = (chart_payload or {}).get('radiacion_mensual', [])
        hsp_mensual = (chart_payload or {}).get('hsp_mensual', [])
        hsp_prom = float((chart_payload or {}).get('hsp_promedio') or 0)
        if radiacion_mensual:
                story.append(_build_radiation_bar_chart(chart_payload, styles))
                story.append(Spacer(1, 8))
                story.append(_build_hsp_bar_chart(chart_payload, styles))
                story.append(Spacer(1, 8))
                radiacion_promedio = sum(radiacion_mensual) / len(radiacion_mensual) if radiacion_mensual else 0
                texto_radiacion = (
                    f"<b>¿Cuál es su potencial de generación solar real?</b> Su zona tiene una radiación solar promedio de <b>{radiacion_promedio:.2f} kWh/m²/día</b> "
                    f"y <b>{hsp_prom:.2f} horas solar pico (HSP)</b> diarias, validando <b>condiciones óptimas para aprovechar la energía del sol</b>. "
                    f"Los datos provienen de PVGIS, la herramienta de modelado de radiación solar más confiable a nivel mundial, respaldada por la Comisión Europea. "
                    f"Su zona está en una <b>ubicación privilegiada</b> con excelente radiación solar durante todo el año, garantizando máximo rendimiento del sistema "
                    f"y máximo ahorro económico. Esta es una oportunidad única para aprovechar un recurso natural confiable, predecible y completamente gratis."
                )
                story.append(Paragraph(texto_radiacion, styles['ExplainText']))
                story.append(Spacer(1, 10))

    # ─── Conditions ───
    if cotizacion.condiciones:
        story.append(Paragraph("Condiciones Comerciales", styles['SectionTitle']))
        for line in cotizacion.condiciones.split('\n'):
            line = line.strip()
            if line:
                story.append(Paragraph(line, styles['ConditionsText']))
        story.append(Spacer(1, 8))

    # ─── Footer note ───
    story.append(HRFlowable(
        width="100%", thickness=1, color=BORDER_COLOR,
        spaceAfter=6, spaceBefore=6,
    ))
    story.append(Paragraph(
        f"Ingeniero Julián A. Castro - {company['name']} - {company['email']} - {company['phone']}.<br/>"
        f"Documento generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} "
        f"por {company['name']}. Este documento no constituye factura de venta.",
        styles['SmallText'],
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer


# ─── PDF Sub-Builders ───

def _build_header(cotizacion, company, styles):
    """Header row with company logo, info, and cotización number."""
    import os

    # ── Logo column (preserve aspect ratio) ──
    logo_path = company.get('logo', '')
    logo_element = None
    if logo_path and os.path.isfile(logo_path):
        try:
            from PIL import Image as PILImg
            with PILImg.open(logo_path) as pil_img:
                orig_w, orig_h = pil_img.size
            max_h = 0.8 * inch
            max_w = 1.1 * inch
            scale = min(max_w / orig_w, max_h / orig_h)
            logo_element = Image(logo_path, width=orig_w * scale, height=orig_h * scale)
            logo_element.hAlign = 'LEFT'
        except Exception:
            logo_element = None

    # ── Company info column ──
    left_content = []
    left_content.append(Paragraph(company['name'], styles['CompanyName']))
    info_parts = []
    if company['nit']:
        info_parts.append(f"NIT: {company['nit']}")
    if company['address']:
        info_parts.append(company['address'])
    if company['phone']:
        info_parts.append(f"Tel: {company['phone']}")
    if company['email']:
        info_parts.append(company['email'])
    if info_parts:
        left_content.append(Paragraph(" | ".join(info_parts), styles['SmallText']))

    # ── Cotización info column ──
    right_content = []
    right_content.append(Paragraph(
        f"<b>COTIZACIÓN</b>", styles['DocTitle']
    ))
    right_content.append(Paragraph(
        f"<b>N° {cotizacion.numero}</b>", styles['RightAligned']
    ))
    right_content.append(Paragraph(
        f"Fecha: {cotizacion.fecha_emision.strftime('%d/%m/%Y')}", styles['RightAligned']
    ))
    right_content.append(Paragraph(
        f"Válida hasta: {cotizacion.fecha_vencimiento.strftime('%d/%m/%Y')}",
        styles['RightAligned'],
    ))

    if logo_element:
        header_table = Table(
            [[logo_element, left_content, right_content]],
            colWidths=[1.2 * inch, 2.8 * inch, 2.8 * inch],
        )
    else:
        header_table = Table(
            [[left_content, right_content]],
            colWidths=[3.8 * inch, 3.0 * inch],
        )
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    return header_table


def _build_info_section(cotizacion, cliente, proyecto, styles):
    """Two-column section with client and project info."""
    client_data = [
        [Paragraph("<b>DATOS DEL CLIENTE</b>", styles['CellText']), ''],
        [Paragraph("Nombre:", styles['SmallText']),
         Paragraph(str(cliente.nombre), styles['CellText'])],
        [Paragraph("Municipio:", styles['SmallText']),
         Paragraph(f"{cliente.ciudad}, {cliente.departamento}", styles['CellText'])],
        [Paragraph("Teléfono:", styles['SmallText']),
         Paragraph(str(cliente.telefono or '—'), styles['CellText'])],
        [Paragraph("Email:", styles['SmallText']),
         Paragraph(str(cliente.email or '—'), styles['CellText'])],
        [Paragraph("Estrato:", styles['SmallText']),
         Paragraph(str(cliente.estrato or '—'), styles['CellText'])],
        [Paragraph("Consumo:", styles['SmallText']),
         Paragraph(f"{cliente.consumo_mensual_kwh} kWh/mes", styles['CellText'])],
    ]

    project_data = [
        [Paragraph("<b>DATOS DEL PROYECTO</b>", styles['CellText']), ''],
        [Paragraph("Código:", styles['SmallText']),
         Paragraph(str(proyecto.codigo), styles['CellText'])],
        [Paragraph("Nombre:", styles['SmallText']),
         Paragraph(str(proyecto.nombre), styles['CellText'])],
        [Paragraph("Tipo:", styles['SmallText']),
         Paragraph(str(proyecto.get_tipo_sistema_display()), styles['CellText'])],
        [Paragraph("Vendedor:", styles['SmallText']),
         Paragraph(
             str(cotizacion.creado_por.get_full_name() if cotizacion.creado_por else '—'),
             styles['CellText'],
         )],
        [Paragraph("Tipo Cliente:", styles['SmallText']),
         Paragraph(str(cotizacion.get_tipo_cliente_display()), styles['CellText'])],
    ]

    client_table = Table(client_data, colWidths=[1.2 * inch, 2.1 * inch])
    client_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fffbeb')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, SOLAR_AMBER),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    project_table = Table(project_data, colWidths=[1.2 * inch, 2.1 * inch])
    project_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fffbeb')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, SOLAR_AMBER),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    wrapper = Table(
        [[client_table, project_table]],
        colWidths=[3.4 * inch, 3.4 * inch],
    )
    wrapper.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    return wrapper


def _build_items_table(cotizacion, styles):
    """Build the items detail table."""
    items = cotizacion.items.select_related('equipo').all()

    header = [
        Paragraph('<b>#</b>', styles['CellText']),
        Paragraph('<b>Equipo</b>', styles['CellText']),
        Paragraph('<b>Categoría</b>', styles['CellText']),
        Paragraph('<b>Cant.</b>', styles['CellText']),
        Paragraph('<b>Precio Unit.</b>', styles['CellText']),
        Paragraph('<b>Desc.</b>', styles['CellText']),
        Paragraph('<b>Subtotal</b>', styles['CellText']),
    ]
    data = [header]

    for i, item in enumerate(items, 1):
        data.append([
            Paragraph(str(i), styles['CellText']),
            Paragraph(
                f"{item.equipo.nombre}<br/>"
                f"<font size='6' color='#9ca3af'>{item.equipo.sku}</font>",
                styles['CellText'],
            ),
            Paragraph(str(item.equipo.get_categoria_display()), styles['CellText']),
            Paragraph(str(item.cantidad), styles['CellText']),
            Paragraph(_fmt_cop(item.precio_unitario), styles['CellText']),
            Paragraph(f"{item.descuento_item}%", styles['CellText']),
            Paragraph(f"<b>{_fmt_cop(item.subtotal)}</b>", styles['CellText']),
        ])

    col_widths = [0.35 * inch, 2.2 * inch, 1.1 * inch, 0.5 * inch, 1.0 * inch, 0.55 * inch, 1.1 * inch]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table_style = [
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), HEADER_TEXT),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
        # Grid
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
    ]

    # Alternate row colors
    for idx in range(1, len(data)):
        if idx % 2 == 0:
            table_style.append(('BACKGROUND', (0, idx), (-1, idx), ROW_ALT))

    table.setStyle(TableStyle(table_style))
    return table


def _build_totals_table(cotizacion, styles):
    """Build the financial totals summary aligned right."""
    rows = []
    if cotizacion.usar_total_manual:
        # Manual total mode – show simplified summary
        rows.append([f'IVA ({cotizacion.iva_porcentaje}%):', _fmt_cop(cotizacion.iva_monto)])
        rows.append(['TOTAL:', _fmt_cop(cotizacion.total)])
    else:
        rows.append(['Subtotal equipos:', _fmt_cop(cotizacion.subtotal)])
        if cotizacion.costo_instalacion:
            rows.append(['Instalación:', _fmt_cop(cotizacion.costo_instalacion)])
        if cotizacion.costo_transporte:
            rows.append(['Transporte:', _fmt_cop(cotizacion.costo_transporte)])
        if cotizacion.descuento_porcentaje:
            rows.append([
                f'Descuento ({cotizacion.descuento_porcentaje}%):',
                f'-{_fmt_cop(cotizacion.descuento_monto)}',
            ])
        rows.append([f'IVA ({cotizacion.iva_porcentaje}%):', _fmt_cop(cotizacion.iva_monto)])
        rows.append(['TOTAL:', _fmt_cop(cotizacion.total)])

    # Convert to Paragraphs
    data = []
    for label, value in rows:
        is_total = label == 'TOTAL:'
        label_style = ParagraphStyle(
            'label', parent=styles['Normal'],
            fontSize=9 if not is_total else 11,
            fontName='Helvetica-Bold' if is_total else 'Helvetica',
            alignment=TA_RIGHT,
            textColor=SOLAR_DARK if is_total else colors.HexColor('#374151'),
        )
        value_style = ParagraphStyle(
            'value', parent=styles['Normal'],
            fontSize=9 if not is_total else 12,
            fontName='Helvetica-Bold',
            alignment=TA_RIGHT,
            textColor=SOLAR_DARK if is_total else colors.HexColor('#111827'),
        )
        data.append([
            Paragraph(label, label_style),
            Paragraph(value, value_style),
        ])

    table = Table(data, colWidths=[2.0 * inch, 1.6 * inch], hAlign='RIGHT')
    table_style = [
        ('BOX', (0, 0), (-1, -1), 0.8, BORDER_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LINEABOVE', (0, -1), (-1, -1), 1.5, SOLAR_AMBER),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fffbeb')),
    ]
    table.setStyle(TableStyle(table_style))
    return table


def _build_executive_summary(cotizacion, proyecto, styles):
    """Commercial summary card with persuasive key indicators."""
    ahorro_anual = float(proyecto.ahorro_mensual or 0) * 12
    costo_total = float(proyecto.costo_total or cotizacion.total or 0)
    roi_anos = (costo_total / ahorro_anual) if ahorro_anual > 0 else 0

    title = Paragraph("Resumen Ejecutivo de la Propuesta", styles['ExecutiveTitle'])
    cards = [
        [
            Paragraph("Inversión", styles['ExecutiveLabel']),
            Paragraph(_fmt_cop(cotizacion.total), styles['ExecutiveMetric']),
        ],
        [
            Paragraph("Ahorro anual estimado", styles['ExecutiveLabel']),
            Paragraph(_fmt_cop(ahorro_anual), styles['ExecutiveMetric']),
        ],
        [
            Paragraph("Retorno estimado", styles['ExecutiveLabel']),
            Paragraph(f"{roi_anos:.1f} años" if roi_anos > 0 else "N/D", styles['ExecutiveMetric']),
        ],
    ]

    metric_table = Table([cards], colWidths=[2.2 * inch, 2.2 * inch, 2.2 * inch])
    metric_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#ffffff')),
        ('BOX', (0, 0), (-1, -1), 0.8, BORDER_COLOR),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))

    mensaje = Paragraph(
        "Esta propuesta prioriza rentabilidad, estabilidad energética y equipos seleccionados desde inventario para asegurar disponibilidad, compatibilidad técnica y una implementación confiable.",
        styles['InventoryNote'],
    )
    wrapper = Table([[title], [metric_table], [mensaje]], colWidths=[6.9 * inch])
    wrapper.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f9fafb')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    return wrapper


def _build_equipment_selection_note(cotizacion, proyecto, styles):
    """Explain inventory-backed equipment selection and technical coherence."""
    items = list(cotizacion.items.select_related('equipo').all())
    if not items:
        return Spacer(1, 1)

    categorias = sorted({item.equipo.get_categoria_display() for item in items})
    total_referencias = len(items)
    texto = (
        f"<b>Selección técnica de equipos desde inventario:</b> se consolidaron <b>{total_referencias} referencias</b> "
        f"disponibles en inventario ({', '.join(categorias)}), priorizando compatibilidad eléctrica y operativa con el dimensionamiento calculado "
        f"para el proyecto <b>{proyecto.nombre}</b>. Esta trazabilidad reduce riesgos de sustitución, mejora tiempos de entrega y mantiene "
        f"coherencia entre ingeniería, costos y ejecución comercial."
    )
    note = Paragraph(texto, styles['InventoryNote'])
    container = Table([[note]], colWidths=[6.9 * inch])
    container.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f9fafb')),
        ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor('#d1d5db')),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    return container


def _build_sizing_summary(proyecto, styles):
    """Table with technical sizing data for the project."""
    data = []
    fields = [
        ('Tipo de sistema', proyecto.get_tipo_sistema_display()),
        ('Potencia del sistema', f"{proyecto.potencia_pico_kwp:.2f} kWp" if proyecto.potencia_pico_kwp else '—'),
        ('Número de paneles', str(proyecto.numero_paneles or '—')),
        ('Generación estimada', f"{proyecto.generacion_mensual_kwh:.0f} kWh/mes" if proyecto.generacion_mensual_kwh else '—'),
        ('Cobertura solar', f"{proyecto.porcentaje_cobertura:.1f}%" if proyecto.porcentaje_cobertura else '—'),
        ('HSP promedio', f"{proyecto.hsp_promedio:.2f} h/día" if proyecto.hsp_promedio else '—'),
        ('Ahorro mensual', _fmt_cop(proyecto.ahorro_mensual) if proyecto.ahorro_mensual else '—'),
        ('ROI estimado', f"{proyecto.roi_anos:.1f} años" if proyecto.roi_anos else '—'),
        ('Costo total del sistema', _fmt_cop(proyecto.costo_total) if proyecto.costo_total else '—'),
    ]

    # Add battery info for off-grid systems
    if proyecto.tipo_sistema == 'off_grid':
        if proyecto.capacidad_baterias_kwh:
            fields.append(('Capacidad baterías', f"{proyecto.capacidad_baterias_kwh:.1f} kWh"))
        if proyecto.autonomia_dias:
            fields.append(('Autonomía', f"{proyecto.autonomia_dias:.1f} días"))

    for label, value in fields:
        data.append([
            Paragraph(f"<b>{label}</b>", styles['CellText']),
            Paragraph(str(value), styles['CellText']),
        ])

    table = Table(data, colWidths=[2.5 * inch, 3.0 * inch])
    table.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fef9c3')),
    ]))
    return table


def _build_cost_pie_chart(chart_payload, _styles=None):
    """Build a ReportLab pie chart for cost distribution from chart payload data."""
    drawing = Drawing(400, 200)
    drawing.hAlign = 'CENTER'

    dist = (chart_payload or {}).get('costo_por_componente', {})
    labels = dist.get('labels') or []
    values = dist.get('values') or []

    if not values or sum(values) == 0:
        return Spacer(1, 1)

    pie = Pie()
    pie.x = 125
    pie.y = 10
    pie.width = 150
    pie.height = 150
    pie.data = values
    pie.labels = [f"{l}\n{_fmt_cop(v)}" for l, v in zip(labels, values)]
    pie.sideLabels = True
    pie.simpleLabels = False
    pie.slices.strokeWidth = 0.5
    pie.slices.strokeColor = colors.white

    # Same color palette as HTML charts (Plotly)
    chart_colors = [
        colors.HexColor('#f59e0b'),
        colors.HexColor('#22c55e'),
        colors.HexColor('#3b82f6'),
        colors.HexColor('#8b5cf6'),
        colors.HexColor('#ef4444'),
        colors.HexColor('#06b6d4'),
        colors.HexColor('#ec4899'),
        colors.HexColor('#84cc16'),
    ]
    for i in range(len(values)):
        pie.slices[i].fillColor = chart_colors[i % len(chart_colors)]
        pie.slices[i].fontSize = 7
        pie.slices[i].fontName = 'Helvetica'

    drawing.add(pie)
    return drawing


def _build_consumption_bar_chart(chart_payload, _styles=None):
    """Build a bar chart comparing current consumption vs solar generation."""
    consumo_data = (chart_payload or {}).get('consumo_comparacion', {})
    consumo = float(consumo_data.get('actual', 0) or 0)
    generacion = float(consumo_data.get('generacion', 0) or 0)
    if consumo <= 0 and generacion <= 0:
        return Spacer(1, 1)

    drawing = Drawing(450, 180)
    drawing.hAlign = 'CENTER'

    # Calculate remaining consumption with solar
    con_solar = max(0, consumo - generacion)
    
    # Manual bar chart drawing for per-bar coloring
    bar_width = 25
    spacing = 80
    y_start = 30
    chart_height = 115
    
    max_val = max(consumo, generacion) * 1.2
    y_scale = chart_height / max_val if max_val > 0 else 1
    
    # Data: consumo (red), generacion (green), con_solar (amber)
    data = [
        (consumo, colors.HexColor('#ef4444'), 'Consumo\nactual'),
        (generacion, colors.HexColor('#22c55e'), 'Generación\nsolar'),
        (con_solar, colors.HexColor('#f59e0b'), 'Consumo\ncon solar'),
    ]
    bar_group_width = ((len(data) - 1) * spacing) + bar_width
    x_start = (450 - bar_group_width) / 2
    
    # Draw bars
    for i, (val, bar_color, label) in enumerate(data):
        x = x_start + i * spacing
        bar_height = val * y_scale
        
        # Draw bar rectangle
        drawing.add(Rect(x, y_start, bar_width, bar_height,
                        fillColor=bar_color, strokeColor=bar_color, strokeWidth=0))
        
        # Draw label below bar
        drawing.add(String(x + bar_width/2, y_start - 20, label,
                          fontSize=7, fontName='Helvetica',
                          fillColor=colors.HexColor('#374151'), textAnchor='middle'))
        
        # Draw value label above bar
        drawing.add(String(x + bar_width/2, y_start + bar_height + 3, f'{val:.0f}',
                          fontSize=6, fontName='Helvetica',
                          fillColor=colors.HexColor('#374151'), textAnchor='middle'))

    # Draw axis line
    drawing.add(Line(x_start, y_start, x_start + bar_group_width, y_start,
                    strokeColor=colors.HexColor('#d1d5db'), strokeWidth=1))
    
    # Draw Y-axis labels
    step = max(1, int(max_val / 5))
    for i in range(0, int(max_val) + 1, step):
        y = y_start + i * y_scale
        drawing.add(String(x_start - 8, y, str(i), fontSize=6,
                          fontName='Helvetica', fillColor=colors.HexColor('#6b7280'),
                          textAnchor='end', textAnchorMode='end'))

    # Y-axis label (vertical)
    y_label = Label()
    y_label.setOrigin(20, y_start + chart_height / 2)
    y_label.angle = 90
    y_label.boxAnchor = 'c'
    y_label.fontName = 'Helvetica'
    y_label.fontSize = 7
    y_label.fillColor = colors.HexColor('#6b7280')
    y_label.setText('kWh/mes')
    drawing.add(y_label)

    # Title
    drawing.add(String(225, 165, 'Consumo vs Generación (kWh/mes)',
                        fontSize=9, fontName='Helvetica-Bold',
                        fillColor=SOLAR_DARK, textAnchor='middle'))

    return drawing


def _build_roi_chart(chart_payload, _styles=None):
    """Build a line chart showing ROI / accumulated savings over 25 years."""
    proyeccion = (chart_payload or {}).get('proyeccion_financiera', {})
    if not proyeccion or 'anos' not in proyeccion:
        return Spacer(1, 1)

    anos = proyeccion['anos']
    ahorro_acumulado = proyeccion['ahorro_acumulado']
    inversion = proyeccion.get('inversion_inicial', 0)

    drawing = Drawing(450, 200)
    drawing.hAlign = 'CENTER'

    chart = LinePlot()
    chart.x = 55
    chart.y = 35
    chart.width = 340
    chart.height = 130

    # Accumulated savings line
    data_line = [(float(a), float(v)) for a, v in zip(anos, ahorro_acumulado)]
    chart.data = [data_line]

    # Green line matching HTML style (#22c55e)
    chart.lines[0].strokeColor = colors.HexColor('#22c55e')
    chart.lines[0].strokeWidth = 2
    chart.lines[0].symbol = makeMarker('Circle')
    chart.lines[0].symbol.size = 2
    chart.lines[0].symbol.fillColor = colors.HexColor('#22c55e')

    chart.xValueAxis.valueMin = 1
    chart.xValueAxis.valueMax = 25
    chart.xValueAxis.valueStep = 5
    chart.xValueAxis.visible = True
    chart.xValueAxis.strokeWidth = 0
    chart.xValueAxis.tickUp = 0
    chart.xValueAxis.tickDown = 0
    chart.xValueAxis.labels.fontSize = 7
    chart.xValueAxis.labels.fontName = 'Helvetica'

    min_val = min(ahorro_acumulado)
    max_val = max(ahorro_acumulado)
    chart.yValueAxis.valueMin = min(0, min_val) * 1.1
    chart.yValueAxis.valueMax = max_val * 1.1
    chart.yValueAxis.labels.fontSize = 6
    chart.yValueAxis.labels.fontName = 'Helvetica'
    chart.yValueAxis.labelTextFormat = lambda x: _fmt_cop(x)

    drawing.add(chart)

    # Zero line (break-even reference) - Red like HTML
    if min_val < 0:
        y_range = max_val * 1.1 - min(0, min_val) * 1.1
        if y_range > 0:
            zero_y = 35 + 130 * (0 - min(0, min_val) * 1.1) / y_range
            drawing.add(Line(55, zero_y, 395, zero_y,
                            strokeColor=colors.HexColor('#ef4444'),
                            strokeWidth=1.5,
                            strokeDashArray=[4, 2]))

    # Title
    drawing.add(String(225, 180, 'Proyección Financiera - Ahorro Acumulado (25 años)',
                        fontSize=9, fontName='Helvetica-Bold',
                        fillColor=SOLAR_DARK, textAnchor='middle'))
    # Y-axis label (vertical)
    y_label = Label()
    y_label.setOrigin(20, 100)
    y_label.angle = 90
    y_label.boxAnchor = 'c'
    y_label.fontName = 'Helvetica'
    y_label.fontSize = 7
    y_label.fillColor = colors.HexColor('#6b7280')
    y_label.setText('COP')
    drawing.add(y_label)

    return drawing


MONTH_NAMES_SHORT = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                     'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']


def _build_radiation_bar_chart(chart_payload, _styles=None):
    """Build a bar chart showing monthly solar radiation from PVGIS with colors."""
    radiacion_mensual = (chart_payload or {}).get('radiacion_mensual', [])
    if not radiacion_mensual or len(radiacion_mensual) == 0:
        return Spacer(1, 1)

    values = [float(r or 0) for r in radiacion_mensual]
    min_rad = min(values)
    max_rad = max(values)

    drawing = Drawing(450, 180)
    drawing.hAlign = 'CENTER'

    # Manual bar chart drawing for per-bar coloring
    bar_width = 20
    spacing = 30
    x_start = 50
    y_start = 30
    chart_height = 115
    chart_width = 350
    
    # Scale values to chart height
    y_scale = chart_height / (max_rad * 1.2) if max_rad > 0 else 1
    
    # Draw bars
    for i, val in enumerate(values):
        x = x_start + i * spacing
        bar_height = val * y_scale
        
        # Choose color: red for min, default amber
        if val == min_rad:
            bar_color = colors.HexColor('#ef4444')
        else:
            bar_color = colors.HexColor('#f59e0b')
        
        # Draw bar rectangle
        drawing.add(Rect(x, y_start, bar_width, bar_height, 
                        fillColor=bar_color, strokeColor=bar_color, strokeWidth=0))
        
        # Draw month label below
        drawing.add(String(x + bar_width/2, y_start - 10, MONTH_NAMES_SHORT[i],
                          fontSize=7, fontName='Helvetica',
                          fillColor=colors.HexColor('#374151'), textAnchor='middle'))
        
        # Draw value label above bar
        drawing.add(String(x + bar_width/2, y_start + bar_height + 3, f'{val:.1f}',
                          fontSize=6, fontName='Helvetica',
                          fillColor=colors.HexColor('#374151'), textAnchor='middle'))

    # Draw axis line
    drawing.add(Line(x_start, y_start, x_start + 350, y_start,
                    strokeColor=colors.HexColor('#d1d5db'), strokeWidth=1))
    
    # Draw Y-axis labels
    for i in range(0, int(max_rad * 1.2) + 1, max(1, int(max_rad * 1.2 / 5))):
        y = y_start + i * y_scale
        drawing.add(String(x_start - 5, y, str(i), fontSize=6, 
                          fontName='Helvetica', fillColor=colors.HexColor('#6b7280'),
                          textAnchor='end'))

    # Title
    drawing.add(String(225, 160, 'Radiación Solar Mensual - PVGIS (kWh/m²)',
                        fontSize=9, fontName='Helvetica-Bold',
                        fillColor=SOLAR_DARK, textAnchor='middle'))

    # X-axis label
    drawing.add(String(225, 10, 'Meses',
                        fontSize=7, fontName='Helvetica',
                        fillColor=colors.HexColor('#6b7280'), textAnchor='middle'))

    # Y-axis label (vertical)
    y_label = Label()
    y_label.setOrigin(20, y_start + chart_height / 2)
    y_label.angle = 90
    y_label.boxAnchor = 'c'
    y_label.fontName = 'Helvetica'
    y_label.fontSize = 7
    y_label.fillColor = colors.HexColor('#6b7280')
    y_label.setText('kWh/m²')
    drawing.add(y_label)

    return drawing


def _build_hsp_bar_chart(chart_payload, _styles=None):
    """Build a bar chart showing monthly HSP with color-coded bars."""
    hsp_values = (chart_payload or {}).get('hsp_mensual', [])
    hsp_promedio = float((chart_payload or {}).get('hsp_promedio') or 0)

    if not hsp_values or len(hsp_values) == 0:
        return Spacer(1, 1)

    hsp_values = [float(h or 0) for h in hsp_values]
    hsp_min = min(hsp_values)
    hsp_max = max(hsp_values)

    drawing = Drawing(450, 180)
    drawing.hAlign = 'CENTER'

    # Manual bar chart drawing for per-bar coloring
    bar_width = 20
    spacing = 30
    x_start = 50
    y_start = 30
    chart_height = 115
    chart_width = 350
    
    # Scale values to chart height
    max_val = hsp_max * 1.3
    y_scale = chart_height / max_val if max_val > 0 else 1
    
    # Draw bars with individual colors
    for i, val in enumerate(hsp_values):
        x = x_start + i * spacing
        bar_height = val * y_scale
        
        # Choose color: red for min, green for max, blue for others
        if val == hsp_min:
            bar_color = colors.HexColor('#ef4444')  # Red
        elif val == hsp_max:
            bar_color = colors.HexColor('#22c55e')  # Green
        else:
            bar_color = colors.HexColor('#3b82f6')  # Blue
        
        # Draw bar rectangle
        drawing.add(Rect(x, y_start, bar_width, bar_height,
                        fillColor=bar_color, strokeColor=bar_color, strokeWidth=0))
        
        # Draw month label below
        drawing.add(String(x + bar_width/2, y_start - 10, MONTH_NAMES_SHORT[i],
                          fontSize=7, fontName='Helvetica',
                          fillColor=colors.HexColor('#374151'), textAnchor='middle'))
        
        # Draw value label above bar
        drawing.add(String(x + bar_width/2, y_start + bar_height + 3, f'{val:.2f}',
                          fontSize=6, fontName='Helvetica',
                          fillColor=colors.HexColor('#374151'), textAnchor='middle'))

    # Draw axis line
    drawing.add(Line(x_start, y_start, x_start + 350, y_start,
                    strokeColor=colors.HexColor('#d1d5db'), strokeWidth=1))
    
    # Draw average line (dashed) in amber
    if hsp_promedio and hsp_promedio > 0:
        avg_line_y = y_start + hsp_promedio * y_scale
        drawing.add(Line(x_start, avg_line_y, x_start + 350, avg_line_y,
                        strokeColor=colors.HexColor('#f59e0b'),
                        strokeWidth=1, strokeDashArray=[4, 2]))
        drawing.add(String(x_start + 355, avg_line_y - 2,
                          f'Prom: {hsp_promedio:.2f}',
                          fontSize=6, fontName='Helvetica',
                          fillColor=colors.HexColor('#f59e0b')))

    # Draw Y-axis labels
    for i in range(0, int(max_val) + 1, max(1, int(max_val / 5))):
        y = y_start + i * y_scale
        drawing.add(String(x_start - 5, y, f'{i:.1f}', fontSize=6,
                          fontName='Helvetica', fillColor=colors.HexColor('#6b7280'),
                          textAnchor='end'))

    # Title
    drawing.add(String(225, 160, 'HSP - Horas Solar Pico (h/día)',
                        fontSize=9, fontName='Helvetica-Bold',
                        fillColor=SOLAR_DARK, textAnchor='middle'))

    # X-axis label
    drawing.add(String(225, 10, 'Meses',
                        fontSize=7, fontName='Helvetica',
                        fillColor=colors.HexColor('#6b7280'), textAnchor='middle'))

    # Y-axis label (vertical)
    y_label = Label()
    y_label.setOrigin(20, y_start + chart_height / 2)
    y_label.angle = 90
    y_label.boxAnchor = 'c'
    y_label.fontName = 'Helvetica'
    y_label.fontSize = 7
    y_label.fillColor = colors.HexColor('#6b7280')
    y_label.setText('h/día')
    drawing.add(y_label)

    return drawing


# ═══════════════════════════════════════════════════════
#  EXCEL GENERATION
# ═══════════════════════════════════════════════════════

def generar_excel_cotizacion(cotizacion):
    """
    Generate an Excel workbook for a cotización.
    Returns an io.BytesIO buffer.
    """
    wb = Workbook()
    company = _company_info()
    proyecto = cotizacion.proyecto
    cliente = proyecto.cliente

    # ─── Styles ───
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
    title_font = Font(name='Calibri', size=14, bold=True, color='92400E')
    subtitle_font = Font(name='Calibri', size=10, color='6B7280')
    bold_font = Font(name='Calibri', size=10, bold=True)
    normal_font = Font(name='Calibri', size=10)
    total_font = Font(name='Calibri', size=12, bold=True, color='92400E')
    total_fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
    border_thin = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )
    cop_format = '#,##0'

    # ══════════════════════════════
    # SHEET 1: Cotización
    # ══════════════════════════════
    ws = wb.active
    ws.title = 'Cotización'
    ws.sheet_properties.tabColor = 'F59E0B'

    # Column widths
    col_widths = [6, 35, 16, 10, 18, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # Company logo (preserve aspect ratio)
    import os
    logo_path = company.get('logo', '')
    logo_inserted = False
    if logo_path and os.path.isfile(logo_path):
        try:
            from PIL import Image as PILImg
            with PILImg.open(logo_path) as pil_img:
                orig_w, orig_h = pil_img.size
            max_h = 60  # pixels
            max_w = 120  # pixels
            scale = min(max_w / orig_w, max_h / orig_h)
            img = OpenpyxlImage(logo_path)
            img.width = int(orig_w * scale)
            img.height = int(orig_h * scale)
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = max(30, img.height * 0.75)
            ws.row_dimensions[2].height = 20
            logo_inserted = True
        except Exception:
            pass

    # Company header (offset to column 2 if logo present)
    logo_col = 2 if logo_inserted else 1
    ws.merge_cells(start_row=row, start_column=logo_col, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=logo_col, value=company['name'])
    cell.font = title_font
    row += 1

    contact_parts = [p for p in [company['nit'], company['address'], company['phone'], company['email']] if p]
    ws.merge_cells(start_row=row, start_column=logo_col, end_row=row, end_column=7)
    ws.cell(row=row, column=logo_col, value=' | '.join(contact_parts)).font = subtitle_font
    row += 2

    # Cotización title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=f'COTIZACIÓN N° {cotizacion.numero}')
    cell.font = Font(name='Calibri', size=14, bold=True, color='92400E')
    cell.alignment = Alignment(horizontal='center')
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1,
                   value=f'Fecha: {cotizacion.fecha_emision.strftime("%d/%m/%Y")}  |  '
                         f'Válida hasta: {cotizacion.fecha_vencimiento.strftime("%d/%m/%Y")}')
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal='center')
    row += 2

    # ─── Client & Project Info ───
    info_labels = [
        ('Cliente:', cliente.nombre, 'Proyecto:', proyecto.codigo),
        ('Municipio:', f'{cliente.ciudad}, {cliente.departamento}', 'Tipo:', proyecto.get_tipo_sistema_display()),
        ('Teléfono:', cliente.telefono or '—', 'Vendedor:',
         cotizacion.creado_por.get_full_name() if cotizacion.creado_por else '—'),
        ('Email:', cliente.email or '—', 'Consumo:', f'{cliente.consumo_mensual_kwh} kWh/mes'),
        ('Estrato:', str(cliente.estrato or '—'), 'Tarifa:', f'{_fmt_cop(cliente.tarifa_electrica)}/kWh'),
    ]

    for label1, val1, label2, val2 in info_labels:
        ws.cell(row=row, column=1, value=label1).font = bold_font
        ws.cell(row=row, column=2, value=val1).font = normal_font
        ws.cell(row=row, column=4, value=label2).font = bold_font
        ws.cell(row=row, column=5, value=val2).font = normal_font
        row += 1

    row += 1

    # ─── Items Header ───
    headers = ['#', 'Equipo', 'Categoría', 'Cant.', 'Precio Unit.', 'Desc. %', 'Subtotal']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    row += 1

    # ─── Items Data ───
    items = cotizacion.items.select_related('equipo').all()
    for i, item in enumerate(items, 1):
        ws.cell(row=row, column=1, value=i).font = normal_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=f'{item.equipo.nombre} ({item.equipo.sku})').font = normal_font
        ws.cell(row=row, column=3, value=item.equipo.get_categoria_display()).font = normal_font
        ws.cell(row=row, column=4, value=item.cantidad).font = normal_font
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')

        cell_precio = ws.cell(row=row, column=5, value=float(item.precio_unitario))
        cell_precio.font = normal_font
        cell_precio.number_format = cop_format
        cell_precio.alignment = Alignment(horizontal='right')

        ws.cell(row=row, column=6, value=f'{item.descuento_item}%').font = normal_font
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')

        cell_sub = ws.cell(row=row, column=7, value=float(item.subtotal))
        cell_sub.font = bold_font
        cell_sub.number_format = cop_format
        cell_sub.alignment = Alignment(horizontal='right')

        # Apply borders
        for col_idx in range(1, 8):
            ws.cell(row=row, column=col_idx).border = border_thin
            if i % 2 == 0:
                ws.cell(row=row, column=col_idx).fill = PatternFill(
                    start_color='FFFBEB', end_color='FFFBEB', fill_type='solid'
                )
        row += 1

    row += 1

    # ─── Totals ───
    totals_data = [
        ('Subtotal:', float(cotizacion.subtotal)),
    ]
    if cotizacion.costo_instalacion:
        totals_data.append(('Instalación:', float(cotizacion.costo_instalacion)))
    if cotizacion.costo_transporte:
        totals_data.append(('Transporte:', float(cotizacion.costo_transporte)))
    if cotizacion.descuento_porcentaje:
        totals_data.append((f'Descuento ({cotizacion.descuento_porcentaje}%):', -float(cotizacion.descuento_monto)))
    totals_data.append((f'IVA ({cotizacion.iva_porcentaje}%):', float(cotizacion.iva_monto)))
    totals_data.append(('TOTAL:', float(cotizacion.total)))

    for label, value in totals_data:
        is_total = label == 'TOTAL:'
        ws.cell(row=row, column=6, value=label).font = total_font if is_total else bold_font
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        cell_val = ws.cell(row=row, column=7, value=value)
        cell_val.font = total_font if is_total else bold_font
        cell_val.number_format = cop_format
        cell_val.alignment = Alignment(horizontal='right')
        if is_total:
            ws.cell(row=row, column=6).fill = total_fill
            cell_val.fill = total_fill
            ws.cell(row=row, column=6).border = Border(top=Side(style='medium', color='F59E0B'))
            cell_val.border = Border(top=Side(style='medium', color='F59E0B'))
        row += 1

    row += 2

    # ─── Conditions ───
    if cotizacion.condiciones:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row=row, column=1, value='Condiciones Comerciales').font = bold_font
        row += 1
        for line in cotizacion.condiciones.split('\n'):
            line = line.strip()
            if line:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                ws.cell(row=row, column=1, value=line).font = Font(name='Calibri', size=9, color='6B7280')
                row += 1

    # ══════════════════════════════
    # SHEET 2: Resumen Técnico
    # ══════════════════════════════
    if proyecto.potencia_pico_kwp:
        ws2 = wb.create_sheet('Resumen Técnico')
        ws2.sheet_properties.tabColor = '10B981'

        for i, w in enumerate([25, 25], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        ws2.merge_cells('A1:B1')
        ws2.cell(row=1, column=1, value='Resumen Técnico del Sistema Solar').font = title_font

        tech_data = [
            ('Tipo de sistema', proyecto.get_tipo_sistema_display()),
            ('Potencia del sistema', f'{proyecto.potencia_pico_kwp:.2f} kWp'),
            ('Número de paneles', str(proyecto.numero_paneles or '—')),
            ('Generación estimada',
             f'{proyecto.generacion_mensual_kwh:.0f} kWh/mes' if proyecto.generacion_mensual_kwh else '—'),
            ('Cobertura solar',
             f'{proyecto.porcentaje_cobertura:.1f}%' if proyecto.porcentaje_cobertura else '—'),
            ('HSP promedio', f'{proyecto.hsp_promedio:.2f} h/día' if proyecto.hsp_promedio else '—'),
            ('Ahorro mensual', _fmt_cop(proyecto.ahorro_mensual) if proyecto.ahorro_mensual else '—'),
            ('ROI estimado', f'{proyecto.roi_anos:.1f} años' if proyecto.roi_anos else '—'),
        ]

        r = 3
        for label, value in tech_data:
            cell_l = ws2.cell(row=r, column=1, value=label)
            cell_l.font = bold_font
            cell_l.fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
            cell_l.border = border_thin

            cell_v = ws2.cell(row=r, column=2, value=value)
            cell_v.font = normal_font
            cell_v.border = border_thin
            r += 1

    # ─── Write to buffer ───
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
